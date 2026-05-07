"""
Export to Excel via openpyxl (lazy import).
"""
from __future__ import annotations

import logging
import os

from schemas import DocumentResult, ExtractedField

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, output_dir: str = "results"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export_document(
        self, result: DocumentResult, output_name: str | None = None
    ) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Document"

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_align = Alignment(horizontal="center")

        ws["A1"] = "Filename"
        ws["B1"] = "Page Count"
        ws["C1"] = "Text Length"
        ws["D1"] = "Extracted Fields"
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws["A2"] = result.filename
        ws["B2"] = result.page_count
        ws["C2"] = len(result.text)
        ws["D2"] = len(result.fields)

        if result.fields:
            fs = wb.create_sheet("Extracted Fields")
            fs["A1"] = "Field Name"
            fs["B1"] = "Value"
            fs["C1"] = "Confidence"
            for col in range(1, 4):
                cell = fs.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            for i, field in enumerate(result.fields, start=2):
                fs.cell(row=i, column=1, value=field.name)
                fs.cell(row=i, column=2, value=field.value[:200])
                fs.cell(row=i, column=3, value=field.confidence)

        ts = wb.create_sheet("Raw Text")
        ts.cell(row=1, column=1, value="Page 1")
        ts.cell(row=2, column=1, value=result.text[:32767])

        filepath = os.path.join(
            self.output_dir,
            output_name or f"{result.filename}_extracted.xlsx",
        )
        wb.save(filepath)
        logger.info("excel exported: %s", filepath)
        return filepath

    def export_fields(
        self, fields: list[ExtractedField], output_name: str = "fields.xlsx"
    ) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Fields"

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws["A1"] = "Field Name"
        ws["B1"] = "Value"
        ws["C1"] = "Confidence"
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        for i, field in enumerate(fields, start=2):
            ws.cell(row=i, column=1, value=field.name)
            ws.cell(row=i, column=2, value=field.value[:200])
            ws.cell(row=i, column=3, value=field.confidence)

        filepath = os.path.join(self.output_dir, output_name)
        wb.save(filepath)
        return filepath
